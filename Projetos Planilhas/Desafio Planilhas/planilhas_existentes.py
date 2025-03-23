import openpyxl

workbook = openpyxl.load_workbook('hockey-players.xlsx')
sheet_player_data = workbook['PlayerData']
# alterando os dados em uma unica celula
print(sheet_player_data['D3'].value)
sheet_player_data['D3'].value = 'AMANDA'
workbook.save('hockey-players.xlsx')

# realizando a varredura em massa
for linha in sheet_player_data.iter_rows(min_row=2,max_row=5,min_col=1,max_col=11):
    print(linha[0].value,linha[1].value,linha[2].value,linha[3].value,linha[4].value,linha[5].value,linha[6].value,linha[7].value,linha[8].value,linha[9].value,linha[10].value)


for coluna in sheet_player_data.iter_cols(min_col=1,max_col=3,min_row=2,max_row=5):
    for celula in coluna:
        print(celula.value)


# Modificando dados em massa
for linha in sheet_player_data.iter_rows(min_row=2):   
    if linha[2].value == 'Canada':
       linha[2].value = 'CANADA'
    if linha[5].value >= 150:
        linha[5].value = 'Divisão A'
workbook.save('hockey-players-nova.xlsx')