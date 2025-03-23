import openpyxl

workbook = openpyxl.load_workbook('hockey-players-nova.xlsx')
print(workbook.sheetnames)

pagina = input('Qual pagina você gostaria de alterar? ')
sheet_usuario = workbook[pagina]
linha_min = int(input('Em qual linha deseja iniciar ? '))
linha_max = int(input('Em qual linha gostaria de finalizar? '))
col_min = int(input('Em qual coluna gostaria de iniciar? '))
col_max = int(input('Em qual coluna gostaria de finalizar? '))

for linha in sheet_usuario.iter_rows(min_row=linha_min,max_row=linha_max,min_col=col_min,max_col=col_max):
    print(linha[0].value,linha[1].value,linha[2].value,linha[3].value,linha[4].value,linha[5].value,linha[6].value,linha[7].value,linha[8].value,linha[9].value,linha[10].value)