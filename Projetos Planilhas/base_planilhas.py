'''
workbook = planilha
sheet = pagina
'''

import openpyxl
import openpyxl.workbook

# criação de planilha
workbook = openpyxl.Workbook()

# mostrar sheets existentes
print(workbook.sheetnames)

# para criar novas sheets
workbook.create_sheet('Planilha 1')
workbook.create_sheet('Planilha 2')
workbook.create_sheet('Planilha 3')

# salvar modificações
#Criando o nome da planilha
workbook.save('endereços.xlsx')

# alterando o nome do sheet
workbook['Planilha 1'].title = 'Planilha 01'
workbook.save('endereços.xlsx')

# apagando um sheet
del workbook['Sheet']

print(workbook.sheetnames)
workbook.save('endereços.xlsx')


# colocaando cabeçalho
# criando uma variavel com o nome da sheet, para facilitar a chamada
sheet_planilha01 = workbook['Planilha 01']
#cabeçlho
sheet_planilha01.append(['rua','cep','casa'])
workbook.save('endereços.xlsx')

# adicionar dados na linha
sheet_planilha01.append(['endereço 1','numero cep 1', 'casa 1'])
sheet_planilha01.append(['endereço 2','numero cep 2', 'casa 2'])
sheet_planilha01.append(['endereço 3','numero cep 3', 'casa 3'])
sheet_planilha01.append(['endereço 4','numero cep 4', 'casa 4'])
workbook.save('endereços.xlsx')

# adicionando dados pelo endereço da cedula
sheet_planilha01['A6'].value = 'endereço 5'
sheet_planilha01['B6'].value = 'numero cep 5'
sheet_planilha01['C6'].value = 'casa 5'


# excluindo linhas
sheet_planilha01.delete_rows(4)
#e para excluir mais de uma , basta adicionar a vergula antes ex: (2,6)
#vai exclir da 2 até a 6
workbook.save('endereços.xlsx')

# excluir colunas
# mesmo processo anterior
sheet_planilha01.delete_cols(3)
workbook.save('endereços.xlsx')

# apagando somente a celula
# 1º nome do sheet  depois nome da celula a ser excluida
del workbook['Planilha 01']['B3']
workbook.save('endereços.xlsx')

'''
base para input

Apos criar a estrutura base da planilha(pagina, linhas e coluna)
cria-se um laço de repetição

continuar == 's'
while continuar =='s':
    rua = input('rua: ')
    cep = input('cep: ')
    casa = input('casa: ')
    sheet__planilha01.append(rua,cep,casa)
    continuar = input('adicionar mais um endereço? (s/n)')

workbook.save('endereços.xlsx')

# se a resposta do usuario for diferente de 's', vaio encerrar o laço e salvar as informações 
'''